import os
import sys
import django
import openpyxl

# Setup Django environment
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "wpc_map.settings")
django.setup()

from facilities.models import Facility, FacilityContact, Program

EXCEL_PATH = "Work Placement Host Orgnisations - Master.xlsx"

def clean(val):
    """Clean and strip excel cell values."""
    if val is None:
        return ""
    return str(val).strip()

def import_contacts():
    """
    Import facility contacts from Excel file.
    Maps Name, Position, and Contact (Phone/Mobile) details.
    """
    print("Starting contact import from Excel...")
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    total_contacts_created = 0
    total_contacts_updated = 0
    
    # Map sheet names to programs
    sheet_program_map = {
        "Individual Support ": "Individual Support",
        "Allied Health ": "Allied Health",
        "Disability": "Disability",
        "ECEC": "ECEC",
    }

    for sheet_name in wb.sheetnames:
        # Only process relevant sheets
        if sheet_name not in sheet_program_map and sheet_name != "MOU (IS and AHA) ":
            if sheet_name in ["Dec gift", "Webinar", "SBES"]:
                continue
        
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        program_name = sheet_program_map.get(sheet_name)
        program = None
        if program_name:
            program, _ = Program.objects.get_or_create(name=program_name)

        # Iterate rows (skip header row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Column mapping based on Excel structure:
            # 0: Organisation
            # 1: Suburb
            # 2: Name (Contact Person)
            # 3: Position
            # 4: Email
            # 5: Address
            # 6: Phone
            # 7: Mobile
            
            org_name = clean(row[0])
            if not org_name:
                continue
            
            contact_name = clean(row[2])
            # Skip if contact name is empty or placeholder
            if not contact_name or contact_name.lower() in ["n/a", "none", "-", "nan"]:
                continue
            
            position = clean(row[3])
            email = clean(row[4])
            # Use Phone or Mobile as the primary contact number
            phone = clean(row[6]) or clean(row[7])
            
            # Match facility name (considering truncation used in main import)
            org_name_truncated = org_name[:120]
            
            # Try exact match first
            facilities = Facility.objects.filter(name__iexact=org_name_truncated)
            
            # If no exact match, try a more flexible search
            if not facilities.exists():
                facilities = Facility.objects.filter(name__icontains=org_name_truncated[:40])
            
            if facilities.exists():
                facility = facilities.first()
                
                # Create or update contact entry
                contact, created = FacilityContact.objects.update_or_create(
                    facility=facility,
                    name=contact_name,
                    defaults={
                        "role": position[:120],
                        "email": email[:254], # EmailField max_length
                        "phone": phone[:50],
                    }
                )
                
                # Link to the program from the current sheet
                if program:
                    contact.programs.add(program)
                
                if created:
                    total_contacts_created += 1
                else:
                    total_contacts_updated += 1
            else:
                # Log if facility not found to help with debugging
                # print(f"  Facility not found in DB: {org_name}")
                pass

    print("\nImport Summary:")
    print(f"  Total Sheets Processed: {len(wb.sheetnames)}")
    print(f"  New Contacts Created: {total_contacts_created}")
    print(f"  Existing Contacts Updated: {total_contacts_updated}")
    print("\nDone!")

if __name__ == "__main__":
    import_contacts()
