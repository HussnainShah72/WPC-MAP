import os
import sys
import django
import openpyxl
import time
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut

# Setup Django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "wpc_map.settings")
django.setup()

from facilities.models import Facility, FacilityContact, Program
from core.wa_suburbs import WA_SUBURB_COORDINATES

EXCEL_PATH = "Work Placement Host Orgnisations - Master.xlsx"

def get_approx_coords(suburb):
    s = str(suburb or "").strip()
    for key, coords in WA_SUBURB_COORDINATES.items():
        if key.lower() == s.lower():
            return coords["latitude"], coords["longitude"]
    return None, None

def clean(val):
    if val is None:
        return ""
    return str(val).strip()

def infer_status(comments, additional):
    text = (clean(comments) + " " + clean(additional)).lower()
    if any(w in text for w in ["active", "confirmed", "approved", "accepted", "ongoing"]):
        return "active_placements"
    if any(w in text for w in ["not available", "no capacity"]):
        return "not_available"
    return "potential"

def import_all_and_contacts():
    print("Importing facilities and contacts from all sheets...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    sheet_program_map = {
        "Individual Support ": "Individual Support",
        "Allied Health ": "Allied Health",
        "Disability": "Disability",
        "ECEC": "ECEC",
    }
    
    fac_count = 0
    contact_count = 0
    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_program_map:
            continue
            
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        program_name = sheet_program_map[sheet_name]
        program, _ = Program.objects.get_or_create(name=program_name)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 0: Org, 1: Suburb, 2: Contact Name, 3: Position, 4: Email, 5: Address, 6: Phone, 7: Mobile
            name = clean(row[0])
            if not name: continue
            
            suburb = clean(row[1])
            contact_name = clean(row[2])
            position = clean(row[3])
            email = clean(row[4])
            address = clean(row[5])
            phone = (clean(row[6]) or clean(row[7]))[:50]
            
            name_truncated = name[:120]
            suburb_truncated = suburb[:120]
            
            lat, lon = get_approx_coords(suburb)
            
            facility, created = Facility.objects.get_or_create(
                name=name_truncated,
                suburb=suburb_truncated,
                defaults={
                    "facility_type": "aged_care" if "aged" in name.lower() or "care" in name.lower() else "other",
                    "latitude": lat,
                    "longitude": lon,
                    "status": infer_status(row[8], row[10]),
                    "phone": phone,
                    "address": address,
                    "quick_notes": f"{clean(row[8])}\n{clean(row[10])}".strip(),
                    "geo_accuracy": "approximate" if lat else "unknown"
                }
            )
            facility.programs.add(program)
            
            if created:
                fac_count += 1
            
            # Import Contact
            if contact_name and contact_name.lower() not in ["n/a", "none", "-", "nan"]:
                contact, c_created = FacilityContact.objects.update_or_create(
                    facility=facility,
                    name=contact_name,
                    defaults={
                        "role": position[:120],
                        "email": email[:254],
                        "phone": phone[:50],
                    }
                )
                contact.programs.add(program)
                if c_created:
                    contact_count += 1
                
    print(f"Import complete! {fac_count} new facilities and {contact_count} new contacts added.")

def geocode_facilities():
    print("Geocoding facilities with missing exact coordinates...")
    geolocator = Nominatim(user_agent="wpc_map_importer_v2")
    
    # Target facilities with unknown or approximate geo
    # Only those with an address
    facilities = Facility.objects.filter(geo_accuracy__in=["unknown", "approximate"]).exclude(address="")
    
    print(f"Found {facilities.count()} facilities to geocode.")
    
    for facility in facilities:
        query = facility.address
        if not query or len(query) < 10:
            query = f"{facility.name}, {facility.suburb}, WA, Australia"
        
        try:
            location = geolocator.geocode(query, timeout=10)
            if location:
                facility.latitude = location.latitude
                facility.longitude = location.longitude
                facility.geo_accuracy = "exact"
                facility.geo_verified = True
                facility.save()
                print(f"  Success: {facility.name}")
            else:
                # Try with name + suburb
                query_fallback = f"{facility.name}, {facility.suburb}, WA, Australia"
                location = geolocator.geocode(query_fallback, timeout=10)
                if location:
                    facility.latitude = location.latitude
                    facility.longitude = location.longitude
                    facility.geo_accuracy = "exact"
                    facility.geo_verified = True
                    facility.save()
                    print(f"  Success (Fallback): {facility.name}")
        except Exception as e:
            pass
        
        time.sleep(1)

if __name__ == "__main__":
    import_all_and_contacts()
    geocode_facilities()
